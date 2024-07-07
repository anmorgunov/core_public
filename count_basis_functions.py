from pathlib import Path

from openpyxl import load_workbook

data_path = Path(__file__).resolve().parent / "Data"

orbToL = {"s": 0, "p": 1, "d": 2, "f": 3, "g": 4, "h": 5}

wb = load_workbook(data_path / "basisSetSizes.xlsx")
ws = wb.active

row = 4
while True:
    contracted = ws["L" + str(row)].value
    if contracted is None:
        break
    bases = contracted.replace("[", "").replace("]", "").split(",")
    n_funcs = 0
    for basis in bases:
        count = int(basis[:-1])
        orb = basis[-1]
        n_funcs += count * (2 * orbToL[orb] + 1)
    ws["O" + str(row)] = contracted
    ws["P" + str(row)] = n_funcs
    row += 1

wb.save(data_path / "basisSetSizes_upd.xlsx")
