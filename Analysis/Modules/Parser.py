import os
from typing import Dict, List, Optional, Set, Union

import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

Number = Union[float, int]
ExperDataType = Dict[str, Number]
BasisDataType = Dict[str, Union[str, Number]]
MoleculeDataType = Dict[str, BasisDataType]
AtomDataType = Dict[str, MoleculeDataType]
AlgoDataType = Dict[str, AtomDataType]

BasisStatsType = Dict[str, Dict[str, Number]]
AtomBasisStatsType = Dict[str, BasisStatsType]
AlgoStatsType = Dict[str, BasisStatsType]
AlgoAtomStatsType = Dict[str, AtomBasisStatsType]


CALC_WB_COLS = {
    "B": "Molecule",
    "C": "Method",
    "D": "Basis",
    "E": "Atom",
    "F": "UHF",
    "G": "MP2",
    "H": "MP3",
    "I": "CCSD",
    "J": "CCSD(T)",
    "K": "Exper.",
    "L": "Frozen",
    "M": "Swapped",
    "N": "T1 for RHF",
    "O": "T1 for UHF(a)",
    "P": "T1 for UHF(b)",
    "Q": "Error",
    "R": "Comments",
    "S": "Custom Notes",
}


def get_next_col(col: str, prefix: Optional[str] = None) -> str:
    if prefix is None:
        prefix = ""
    strings = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
    if not col:
        return "A"
    if len(col) == 1:
        if col == "Z":
            return get_next_col(prefix, prefix=None) + get_next_col("", prefix=None)
        else:
            return prefix + strings[strings.index(col) + 1]
    else:
        return get_next_col(col[1:], prefix=prefix + col[0])


class ParsedData:
    """
    This object takes an excel file with parsed data and updates it with new results. The advantage of this approach is that any custom comments left in column L in CEBE_Data file are preserved.
    """

    specialBases = {
        "ccX-DZ",
        "ccX-TZ",
        "ccX-QZ",
        "ccX-5Z",
        "pcX-1",
        "pcX-2",
        "pcX-3",
        "pcX-4",
    }

    def __init__(
        self,
        experimental_wb: str,
        calculations_folder: str,
        algorithms: List[str],
        calculations_wb: str,
        do_average_experimental: bool = True,
    ):
        """
        Creates a data object. Main reason for using a class is to avoid having to pass different data objects from function to funtion
        """
        self.experWB = load_workbook(experimental_wb)
        self.readWB = load_workbook(calculations_wb)
        self.savePath = calculations_wb
        self.calcPath = calculations_folder

        self.saveWB = Workbook()

        self.molToExper: ExperDataType = {}
        self.doAverageExperimental = do_average_experimental

        self.algorithms = algorithms
        self.algoToData = {}

        self.debug = False

        self.font = Font(name="Helvetica", size=12)
        self.cols = "B C D E F G H I J K L M N O P Q R S"

    def _parse_calculations(self):
        """
        Parse data from folders listed in self.algorithmToFolders and store it in a molToData dictionary

        molToData maps molecule (str) to basis (D/T, also str) to a dictionary:
            \{
                "atom": atom that is excited
                "error": error if any
                "UHF": result of UHF calculation
                "MP2": result of MP2 calculation
                "MP3": result of MP3 calculation
                "CCSD": result of CCSD calculation
            \}

        molToData is saved as a value of self.algoToData with relevant key

        P.S. This is already implemented to work with different algorithms (mom, sgm, coreless). Just use a different key in algorithmToFolders
        """
        for algorithm in self.algorithms:
            algoPath = os.path.join(self.calcPath, algorithm)
            atoms = [
                item
                for item in os.listdir(algoPath)
                if os.path.isdir(os.path.join(algoPath, item)) and item != ".DS_Store"
            ]
            if self.debug:
                print(f"{atoms=}")
            for atom in atoms:
                atomToData: AtomDataType = self.algoToData.setdefault(
                    algorithm, {}
                ).setdefault(atom, {})

                if self.debug:
                    print(f"{self.algoToData=}, {atomToData=}")
                atomPath = os.path.join(algoPath, atom)
                molecules = [
                    molecule
                    for molecule in os.listdir(atomPath)
                    if os.path.isdir(os.path.join(atomPath, molecule))
                ]
                if self.debug:
                    print(f"{molecules=}")
                for molecule in molecules:
                    molPath = os.path.join(atomPath, molecule)
                    bases = [
                        basis
                        for basis in os.listdir(molPath)
                        if os.path.isdir(os.path.join(molPath, basis))
                    ]
                    if self.debug:
                        print(f"{bases=}")
                    for basis in bases:
                        basisToData: BasisDataType = atomToData.setdefault(
                            molecule, {}
                        ).setdefault(basis, {})
                        basPath = os.path.join(molPath, basis)
                        files = os.listdir(basPath)
                        if self.debug:
                            print(f"{files=}")

                        # Case 1. Calculation was unsuccesful and did not terminate normally
                        if algorithm == "gracemom":
                            suffix = "mom"
                        else:
                            suffix = "mom"
                        if f"CEBE_{suffix}.txt" not in files:
                            basisToData["error"] = "No CEBE file"
                            err = os.path.join(basPath, "sbatch.err")
                            if "sbatch.err" not in files:
                                basisToData["detailed"] = "empty?"
                                continue
                            errFile = open(err, "r").readlines()
                            if errFile:
                                basisToData["detailed"] = errFile[-1].split("\n")[0]
                            else:
                                basisToData["detailed"] = "empty?"
                        # Case 2. Calculation terminated normally
                        else:
                            result = os.path.join(basPath, f"CEBE_{suffix}.txt")
                            lines = open(result, "r").readlines()
                            for rline in lines:
                                line = rline.split("\n")[0]
                                if ":" in line:
                                    key, val = rline.split("\n")[0].split(":")
                                    basisToData[key] = val
                                    continue
                                if "=" not in line:
                                    continue
                                rmethod, rvalue = line.split(" = ")
                                value = float(rvalue.split(" ")[0])
                                posthf_method = rmethod.split(" ")[1][1:-1]
                                basisToData[posthf_method] = value

    def _parse_experimental(self):
        """
        Parse data from experimental_wb file and create a dictionary mapping molecule name (which is used to denote file names) to experimental values
        """
        ws = self.experWB["Sheet1"]
        data = {}
        row = 2
        while True:
            if all([ws[col + str(row)].value is None for col in ("A", "B", "C")]):
                # empty row reached = end of file
                break
            if all([ws[col + str(row)].value is not None for col in ("A", "B", "C")]):
                # only parse rows with all three columns filled
                if self.doAverageExperimental:
                    exprs = [ws["B" + str(row)].value]
                    col = "C"
                    while True:
                        col = get_next_col(col)
                        if (val := ws[col + str(row)].value) is None:
                            break
                        exprs.append(val)
                    assert all([type(expr) in {float, int} for expr in exprs])
                    expr = np.mean(exprs)
                else:
                    expr = ws["B" + str(row)].value
                fname = ws["C" + str(row)].value
                data[fname] = expr
            row += 1
        self.molToExper = data

    def _calculate_mp25(self):
        for atomData in self.algoToData.values():
            for molData in atomData.values():
                for basData in molData.values():
                    for data in basData.values():
                        if "MP2" in data and "MP3" in data:
                            data["MP2.5"] = (data["MP2"] + data["MP3"]) / 2

    def _update_calculation_wb(self):
        """
        Write the contents of self.algoToData from parse_results to a CEBE_Data file, preserving any comments in column L (currently not used)
        """
        readWS = self.readWB["Sheet"]
        ws = self.saveWB["Sheet"]

        r = 3
        for col, name in CALC_WB_COLS.items():
            ws[col + str(r)] = name

        for col in self.cols.split(" "):
            ws[col + str(r)].font = self.font
        r += 1
        for algorithm, atomData in self.algoToData.items():
            for atom, molData in atomData.items():
                for molecule, basToData in molData.items():
                    for basis, data in basToData.items():
                        ws["B" + str(r)] = molecule
                        ws["C" + str(r)] = algorithm
                        ws["D" + str(r)] = basis
                        ws["E" + str(r)] = atom
                        if molecule in self.molToExper:
                            ws["K" + str(r)] = self.molToExper[molecule]
                        if "error" in data:
                            ws["Q" + str(r)] = data["error"]
                            ws["R" + str(r)] = data["detailed"]
                        else:
                            if "UHF" in data:
                                ws["F" + str(r)] = data["UHF"]
                            if "MP2" in data:
                                ws["G" + str(r)] = data["MP2"]
                            if "MP3" in data:
                                ws["H" + str(r)] = data["MP3"]
                            if "CCSD" in data:
                                ws["I" + str(r)] = data["CCSD"]
                            if "CCSD(T)" in data:
                                ws["J" + str(r)] = data["CCSD(T)"]
                            if "CCSDT" in data:
                                ws["J" + str(r)] = data["CCSDT"]
                            if "Frozen orbitals" in data:
                                ws["L" + str(r)] = data["Frozen orbitals"]
                            if "Swapped orbitals" in data:
                                ws["M" + str(r)] = data["Swapped orbitals"]
                            if "T1 for RHF" in data:
                                ws["N" + str(r)] = round(float(data["T1 for RHF"]), 4)
                            if "T1 for UHF(a)" in data:
                                ws["O" + str(r)] = round(
                                    float(data["T1 for UHF(a)"]), 4
                                )
                            if "T1 for UHF(b)" in data:
                                ws["P" + str(r)] = round(
                                    float(data["T1 for UHF(b)"]), 4
                                )
                        ws["S" + str(r)] = readWS["S" + str(r)].value
                        for col in self.cols.split(" "):
                            ws[col + str(r)].font = self.font

                        r += 1

        self.saveWB.save(self.savePath)

    def extract_molecules(self, mols: Set[str], save_path: str):
        """ """
        new_wb = Workbook()
        new_ws = new_wb["Sheet"]

        wb = load_workbook(self.savePath)
        ws = wb["Sheet"]

        row = 3
        for col, name in CALC_WB_COLS.items():
            new_ws[col + str(row)] = name

        for col in self.cols.split(" "):
            new_ws[col + str(3)].font = self.font
        row, new_row = 4, 4
        while True:
            if (mol := ws["B" + str(row)].value) is None:
                break
            if mol in mols:
                new_ws["B" + str(new_row)] = mol
                for col in self.cols.split():
                    new_ws[col + str(new_row)] = ws[col + str(row)].value
                for col in self.cols.split(" "):
                    new_ws[col + str(new_row)].font = self.font
                new_row += 1

            row += 1
        new_wb.save(save_path)

    def filter_data_by_molecules(
        self, algoToData: AlgoDataType, atomToMols: Dict[str, Set[str]]
    ) -> AlgoDataType:
        filtered_algoToData = {}
        for algorithm, atomData in algoToData.items():
            for atom, molData in atomData.items():
                if atom not in atomToMols:
                    continue
                for molecule, basData in molData.items():
                    if molecule in atomToMols[atom]:
                        filtered_algoToData.setdefault(algorithm, {}).setdefault(
                            atom, {}
                        )[molecule] = basData
                        filtered_algoToData[algorithm][atom][molecule] = basData
        return filtered_algoToData

    def filter_by_presence_of_experimental(
        self, algoToData: AlgoDataType
    ) -> AlgoDataType:
        filtered_algoToData = {}
        for algorithm, atomData in algoToData.items():
            for atom, molData in atomData.items():
                for molecule, basData in molData.items():
                    if molecule in self.molToExper:
                        filtered_algoToData.setdefault(algorithm, {}).setdefault(
                            atom, {}
                        )[molecule] = basData
        return filtered_algoToData

    def calculate_errors(self, algoToData: AlgoDataType):
        valid_keys = {"UHF", "MP2", "MP2.5", "MP3", "CCSD", "CCSD(T)"}
        algoToError = {}
        for algorithm, atomData in algoToData.items():
            for atom, molData in atomData.items():
                for molecule, basData in molData.items():
                    if molecule not in self.molToExper:
                        continue
                    for bas, methodData in basData.items():
                        for key, value in methodData.items():
                            if key not in valid_keys:
                                continue
                            error = value - self.molToExper[molecule]
                            if error > 1 and bas in self.specialBases:
                                if self.debug:
                                    print(
                                        f"Large error for {algorithm} {atom} {molecule} {bas} {key}: {error}"
                                    )
                            algoToError.setdefault(algorithm, {}).setdefault(
                                atom, {}
                            ).setdefault(molecule, {}).setdefault(bas, {})[key] = error
        self.algoToError = algoToError

    def calculate_series_statistics(self):
        assert hasattr(self, "algoToError"), "You must call calculate_errors first"
        algoToAtomErrors = {}
        for algorithm, atomData in self.algoToError.items():
            for atom, molData in atomData.items():
                for basData in molData.values():
                    for bas, methodData in basData.items():
                        for method, value in methodData.items():
                            algoToAtomErrors.setdefault(algorithm, {}).setdefault(
                                atom, {}
                            ).setdefault(bas, {}).setdefault(method, []).append(value)
        algoToAtomStats: AlgoAtomStatsType = {}
        for algorithm, atomData in algoToAtomErrors.items():
            for atom, basData in atomData.items():
                for bas, methodData in basData.items():
                    for method, error_list in methodData.items():
                        errors = np.array(error_list)
                        abs_errs = np.abs(errors)
                        algoToAtomStats.setdefault(algorithm, {}).setdefault(
                            atom, {}
                        ).setdefault(bas, {})[method] = {
                            "MSE": np.mean(errors),
                            "MAE": np.mean(abs_errs),
                            "MedAE": np.median(abs_errs),
                            "MaxAE": np.max(abs_errs),
                            "STD(AE)": np.std(abs_errs),
                            "n": len(errors),
                        }
        self.algoToAtomStats = algoToAtomStats

    def calculate_overall_statistics(self):
        assert hasattr(self, "algoToError"), "You must call calculate_errors first"
        algoToErrors = {}
        for algorithm, atomData in self.algoToError.items():
            for molData in atomData.values():
                for basData in molData.values():
                    for bas, methodData in basData.items():
                        for method, value in methodData.items():
                            algoToErrors.setdefault(algorithm, {}).setdefault(
                                bas, {}
                            ).setdefault(method, []).append(value)
        algoToStats: AlgoStatsType = {}
        for algorithm, basData in algoToErrors.items():
            for bas, methodData in basData.items():
                for method, error_list in methodData.items():
                    errors = np.array(error_list)
                    abs_errs = np.abs(errors)
                    algoToStats.setdefault(algorithm, {}).setdefault(bas, {})[
                        method
                    ] = {
                        "MSE": np.mean(errors),
                        "MAE": np.mean(abs_errs),
                        "MedAE": np.median(abs_errs),
                        "MaxAE": np.max(abs_errs),
                        "STD(AE)": np.std(abs_errs),
                        "n": len(errors),
                    }
        self.algoToStats = algoToStats

    def main(self, save: bool = True):
        self._parse_experimental()
        self._parse_calculations()
        self._calculate_mp25()
        if save:
            self._update_calculation_wb()


if __name__ == "__main__":
    # perform_parsing()
    pass
