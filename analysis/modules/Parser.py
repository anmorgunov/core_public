from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font
import os
import numpy as np
from typing import List, Set, Union, Dict, Optional


Number = Union[float, int]
ExperDataType = Dict[str, Number]
BasisDataType = Dict[str, Union[str, Number]]
MoleculeDataType = Dict[str, BasisDataType]
AtomDataType = Dict[str, MoleculeDataType]
MethodDataType = Dict[str, AtomDataType]


CALC_WB_COLS = {
    "B": "Molecule",
    "C": "Method",
    "D": "Basis",
    "E": "Atom",
    "F": "UHF",
    "G": "MP2",
    "H": "MP3",
    "I": "CCSD",
    "J": "CCSD(T)",
    "K": "Exper.",
    "L": "Frozen",
    "M": "Swapped",
    "N": "T1 for RHF",
    "O": "T1 for UHF(a)",
    "P": "T1 for UHF(b)",
    "Q": "Error",
    "R": "Comments",
    "S": "Custom Notes",
}


def get_next_col(col:str, prefix:Optional[str]=None)->str:
    if prefix is None:
        prefix = ""
    strings = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
    if not col:
        return "A"
    if len(col) == 1:
        if col == "Z":
            return get_next_col(prefix, prefix=None) + get_next_col("", prefix=None)
        else:
            return prefix + strings[strings.index(col) + 1]
    else:
        return get_next_col(col[1:], prefix=prefix + col[0])


class ParsedData:
    """
    This object takes an excel file with parsed data and updates it with new results. The advantage of this approach is that any custom comments left in column L in CEBE_Data file are preserved.
    """

    def __init__(
        self,
        experimental_wb: str,
        calculations_folder: str,
        methods: List[str],
        calculations_wb: str,
        do_average_experimental: bool = True,
    ):
        """
        Creates a data object. Main reason for using a class is to avoid having to pass different data objects from function to funtion
        """
        self.experWB = load_workbook(experimental_wb)
        self.readWB = load_workbook(calculations_wb)
        self.savePath = calculations_wb
        self.calcPath = calculations_folder

        self.saveWB = Workbook()

        self.molToExper:ExperDataType = {}
        self.doAverageExperimental = do_average_experimental

        self.methods = methods
        self.methodToData = {}

        self.debug = False

        self.font = Font(name="Helvetica", size=12)
        self.cols = "B C D E F G H I J K L M N O P Q R S"

    def _parse_calculations(self):
        """
        Parse data from folders listed in self.methodToFolders and store it in a molToData dictionary

        molToData maps molecule (str) to basis (D/T, also str) to a dictionary:
            \{
                "atom": atom that is excited
                "error": error if any
                "UHF": result of UHF calculation
                "MP2": result of MP2 calculation
                "MP3": result of MP3 calculation
                "CCSD": result of CCSD calculation
            \}

        molToData is saved as a value of self.methodToData with relevant key

        P.S. This is already implemented to work with different methods (mom, sgm, coreless). Just use a different key in methodToFolders
        """
        for method in self.methods:
            methodPath = os.path.join(self.calcPath, method)
            atoms = [
                item
                for item in os.listdir(methodPath)
                if os.path.isdir(os.path.join(methodPath, item)) and item != ".DS_Store"
            ]
            if self.debug:
                print(f"{atoms=}")
            for atom in atoms:
                atomToData:AtomDataType = self.methodToData.setdefault(method, {}).setdefault(
                    atom, {}
                )

                if self.debug:
                    print(f"{self.methodToData=}, {atomToData=}")
                atomPath = os.path.join(methodPath, atom)
                molecules = [
                    molecule
                    for molecule in os.listdir(atomPath)
                    if os.path.isdir(os.path.join(atomPath, molecule))
                ]
                if self.debug:
                    print(f"{molecules=}")
                for molecule in molecules:
                    molPath = os.path.join(atomPath, molecule)
                    bases = [
                        basis
                        for basis in os.listdir(molPath)
                        if os.path.isdir(os.path.join(molPath, basis))
                    ]
                    if self.debug:
                        print(f"{bases=}")
                    for basis in bases:
                        basisToData:BasisDataType = atomToData.setdefault(molecule, {}).setdefault(
                            basis, {}
                        )
                        basPath = os.path.join(molPath, basis)
                        files = os.listdir(basPath)
                        if self.debug:
                            print(f"{files=}")

                        # Case 1. Calculation was unsuccesful and did not terminate normally
                        if f"CEBE_{method}.txt" not in files:
                            basisToData["error"] = "No CEBE file"
                            err = os.path.join(basPath, "sbatch.err")
                            errFile = open(err, "r").readlines()
                            if errFile:
                                basisToData["detailed"] = errFile[-1].split("\n")[0]
                            else:
                                basisToData["detailed"] = "empty?"
                        # Case 2. Calculation terminated normally
                        else:
                            result = os.path.join(basPath, f"CEBE_{method}.txt")
                            lines = open(result, "r").readlines()
                            for rline in lines:
                                line = rline.split("\n")[0]
                                if ":" in line:
                                    key, val = rline.split("\n")[0].split(":")
                                    basisToData[key] = val
                                    continue
                                if "=" not in line:
                                    continue
                                rmethod, rvalue = line.split(" = ")
                                value = float(rvalue.split(" ")[0])
                                posthf_method = rmethod.split(" ")[1][1:-1]
                                basisToData[posthf_method] = value

    def _parse_experimental(self):
        """
        Parse data from experimental_wb file and create a dictionary mapping molecule name (which is used to denote file names) to experimental values
        """
        ws = self.experWB["Sheet1"]
        data = {}
        row = 2
        while True:
            if all([ws[col + str(row)].value is None for col in ("A", "B", "C")]):
                # empty row reached = end of file
                break
            if all([ws[col + str(row)].value is not None for col in ("A", "B", "C")]):
                # only parse rows with all three columns filled
                if self.doAverageExperimental:
                    exprs = [ws["B" + str(row)].value]
                    col = "C"
                    while True:
                        col = get_next_col(col)
                        if (val := ws[col + str(row)].value) is None:
                            break
                        exprs.append(val)
                    assert all([type(expr) in {float, int} for expr in exprs])
                    expr = np.mean(exprs)
                else:
                    expr = ws["B" + str(row)].value
                fname = ws["C" + str(row)].value
                data[fname] = expr
            row += 1
        self.molToExper = data

    def _calculate_mp25(self):
        for atomData in self.methodToData.values():
            for molData in atomData.values():
                for basData in molData.values():
                    for data in basData.values():
                        if "MP2" in data and "MP3" in data:
                            data["MP2.5"] = (data["MP2"] + data["MP3"]) / 2

    def _update_calculation_wb(self):
        """
        Write the contents of self.methodToData from parse_results to a CEBE_Data file, preserving any comments in column L (currently not used)
        """
        readWS = self.readWB["Sheet"]
        ws = self.saveWB["Sheet"]

        r = 3
        for col, name in CALC_WB_COLS.items():
            ws[col+str(r)] = name

        for col in self.cols.split(" "):
            ws[col + str(r)].font = self.font
        r += 1
        for method, atomData in self.methodToData.items():
            for atom, molData in atomData.items():
                for molecule, basToData in molData.items():
                    for basis, data in basToData.items():
                        ws["B" + str(r)] = molecule
                        ws["C" + str(r)] = method
                        ws["D" + str(r)] = basis
                        ws["E" + str(r)] = atom
                        if molecule in self.molToExper:
                            ws["K" + str(r)] = self.molToExper[molecule]
                        if "error" in data:
                            ws["Q" + str(r)] = data["error"]
                            ws["R" + str(r)] = data["detailed"]
                        else:
                            if "UHF" in data:
                                ws["F" + str(r)] = data["UHF"]
                            if "MP2" in data:
                                ws["G" + str(r)] = data["MP2"]
                            if "MP3" in data:
                                ws["H" + str(r)] = data["MP3"]
                            if "CCSD" in data:
                                ws["I" + str(r)] = data["CCSD"]
                            if "CCSD(T)" in data:
                                ws["J" + str(r)] = data["CCSD(T)"]
                            if "CCSDT" in data:
                                ws["J" + str(r)] = data["CCSDT"]
                            if "Frozen orbitals" in data:
                                ws["L" + str(r)] = data["Frozen orbitals"]
                            if "Swapped orbitals" in data:
                                ws["M" + str(r)] = data["Swapped orbitals"]
                            if "T1 for RHF" in data:
                                ws["N" + str(r)] = round(float(data["T1 for RHF"]), 4)
                            if "T1 for UHF(a)" in data:
                                ws["O" + str(r)] = round(
                                    float(data["T1 for UHF(a)"]), 4
                                )
                            if "T1 for UHF(b)" in data:
                                ws["P" + str(r)] = round(
                                    float(data["T1 for UHF(b)"]), 4
                                )
                        ws["S" + str(r)] = readWS["S" + str(r)].value
                        for col in self.cols.split(" "):
                            ws[col + str(r)].font = self.font

                        r += 1

        self.saveWB.save(self.savePath)

    def extract_molecules(self, mols:Set[str], save_path:str):
        """
        """
        new_wb = Workbook()
        new_ws = new_wb["Sheet"]
        
        wb = load_workbook(self.savePath)
        ws = wb["Sheet"]

        row = 3
        for col, name in CALC_WB_COLS.items():
            new_ws[col + str(row)] = name

        for col in self.cols.split(" "):
            new_ws[col + str(3)].font = self.font
        row, new_row = 4, 4
        while True:
            if (mol := ws["B" + str(row)].value) is None:
                break
            if mol in mols:
                new_ws["B" + str(new_row)] = mol
                for col in self.cols.split():
                    new_ws[col + str(new_row)] = ws[col + str(row)].value
                for col in self.cols.split(" "):
                    new_ws[col + str(new_row)].font = self.font
                new_row += 1

            row += 1
        new_wb.save(save_path)

    def filter_data_by_molecules(self, methodToData:MethodDataType, atomToMols:Dict[str, Set[str]]) -> MethodDataType:
        filtered_methodToData = {}
        for method, atomData in methodToData.items():
            for atom, molData in atomData.items():
                for molecule, basData in molData.items():
                    if molecule in atomToMols[atom]:
                        filtered_methodToData.setdefault(method, {}).setdefault(atom, {})[molecule] = basData
                        filtered_methodToData[method][atom][molecule] = basData
        return filtered_methodToData

    def filter_by_presence_of_experimental(self, methodToData:MethodDataType) -> MethodDataType:
        filtered_methodToData = {}
        for method, atomData in methodToData.items():
            for atom, molData in atomData.items():
                for molecule, basData in molData.items():
                    if molecule in self.molToExper:
                        filtered_methodToData.setdefault(method, {}).setdefault(atom, {})[molecule] = basData
        return filtered_methodToData

    def main(self, save: bool = True):
        self._parse_experimental()
        self._parse_calculations()
        self._calculate_mp25()
        if save:
            self._update_calculation_wb()

if __name__ == "__main__":
    # perform_parsing()
    pass
