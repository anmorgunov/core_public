import re
from pathlib import Path
from typing import Dict, List, Literal, Optional, Set, TypedDict, cast

import numpy as np
import numpy.typing as npt
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

HA_TO_EV = 27.211399
# fmt:off
allowed_extra_keys = {"Frozen orbitals", "Swapped orbitals", "T1 for RHF", "T1 for UHF(a)", "T1 for UHF(b)", "error", "detailed"}
allowed_methods = {"HF", "UHF", "MP2", "MP3", "CCSD", "CCSD(T)", "CCSDT"}
LitKeyType = Literal["Frozen orbitals", "Swapped orbitals", "T1 for RHF", "T1 for UHF(a)", "T1 for UHF(b)", "UHF", "MP2", "MP3", "CCSD", "CCSD(T)", "CCSDT", "error", "detailed"]
# fmt:on
DeltaEnergyType = TypedDict(
    "DeltaEnergyType",
    {
        "Frozen orbitals": str,
        "Swapped orbitals": str,
        "UHF": float,
        "MP2": float,
        "MP2.5": float,
        "MP3": float,
        "CCSD": float,
        "CCSD(T)": float,
        "CCSDT": float,
        "T1 for RHF": str,
        "T1 for UHF(a)": str,
        "T1 for UHF(b)": str,
        "error": str,
        "detailed": str,
    },
    total=False,
)

ExperDataType = Dict[str, float]
# DataPointType = Dict[str, Union[str, float]]
BasisDeltaType = Dict[str, DeltaEnergyType]
MoleculeDeltaType = Dict[str, BasisDeltaType]
AtomDeltaType = Dict[str, MoleculeDeltaType]
AlgoDeltaType = Dict[str, AtomDeltaType]

MethodErrorType = Dict[str, float]
BasisErrorType = Dict[str, MethodErrorType]
MoleculeErrorType = Dict[str, BasisErrorType]
AtomErrorType = Dict[str, MoleculeErrorType]
AlgoErrorType = Dict[str, AtomErrorType]

MethodErrorsType = Dict[str, List[float]]
BasisErrorsType = Dict[str, MethodErrorsType]
AlgoErrorsType = Dict[str, BasisErrorsType]
AtomErrorsType = Dict[str, BasisErrorsType]
AlgoAtomErrorsType = Dict[str, AtomErrorsType]

# fmt:off
StatsKeyType = Literal["MSE", "MAE", "MedAE", "MaxAE", "STD(AE)", "n", "errors", "abs_errors"]
StatsType = TypedDict('StatsType', {"MSE":float, "MAE":float, "MedAE": float, "MaxAE": float, "STD(AE)":float, "n": int, "errors": npt.NDArray[np.float64], "abs_errors":npt.NDArray[np.float64]})
# fmt:on
# StatsType = Dict[str, Union[float, int]]
MethodStatsType = Dict[str, StatsType]
BasisStatsType = Dict[str, MethodStatsType]
AlgoStatsType = Dict[str, BasisStatsType]
AtomBasisStatsType = Dict[str, BasisStatsType]
AlgoAtomStatsType = Dict[str, AtomBasisStatsType]


class MethodResultType(TypedDict):
    tot: float
    corr: float


# fmt:off
CALC_WB_COLS = {"B": "Molecule","C": "Method","D": "Basis","E": "Atom","F": "UHF","G": "MP2","H": "MP3","I": "CCSD","J": "CCSD(T)","K": "Exper.","L": "Frozen","M": "Swapped","N": "T1 for RHF","O": "T1 for UHF(a)","P": "T1 for UHF(b)","Q": "Error","R": "Comments","S": "Custom Notes"}
StateResults = TypedDict("StateResults", {"RHF": float, "UHF": float, "RHF-MP2": MethodResultType, "RHF-CCSD": MethodResultType, "RHF-CCSD(T)": MethodResultType, "UHF-MP2": MethodResultType, "UHF-CCSD": MethodResultType, "UHF-CCSD(T)": MethodResultType})
MethodKeyType = Literal["RHF", "UHF", "RHF-MP2", "RHF-CCSD", "RHF-CCSD(T)", "UHF-MP2", "UHF-CCSD", "UHF-CCSD(T)"]
# fmt:on
BasisEnergyType = Dict[str, StateResults]
MoleculeEnergyType = Dict[str, BasisEnergyType]
AtomEnergyType = Dict[str, MoleculeEnergyType]
AlgoEnergyType = Dict[str, AtomEnergyType]


def parse_pyscf_output(file_path: str | Path) -> StateResults:
    with open(file_path, "r") as file:
        content = file.read()

    results: StateResults = {
        "RHF": 0.0,
        "UHF": 0.0,
        "RHF-MP2": {"tot": 0.0, "corr": 0.0},
        "RHF-CCSD": {"tot": 0.0, "corr": 0.0},
        "RHF-CCSD(T)": {"tot": 0.0, "corr": 0.0},
        "UHF-MP2": {"tot": 0.0, "corr": 0.0},
        "UHF-CCSD": {"tot": 0.0, "corr": 0.0},
        "UHF-CCSD(T)": {"tot": 0.0, "corr": 0.0},
    }

    patterns: Dict[MethodKeyType, str] = {
        "RHF": r">>> Running RHF <<<.*?final Etot\s+:\s+(-?\d+\.\d+)",
        "UHF": r">>> Running UHF <<<.*?final Etot\s+:\s+(-?\d+\.\d+)",
        "RHF-MP2": r">>> Running RHF-MP2 <<<.*?MP2 correlation energy \(RHF\) : (-?\d+\.\d+).*?MP2 total energy\s+\(RHF\) : (-?\d+\.\d+)",
        "UHF-MP2": r">>> Running UHF-MP2 <<<.*?MP2 correlation energy \(UHF\) : (-?\d+\.\d+).*?MP2 total energy\s+\(UHF\) : (-?\d+\.\d+)",
        "RHF-CCSD": r">>> Running RHF-CCSD <<<.*?CCSD correlation energy\s+\(RHF\) : (-?\d+\.\d+).*?CCSD total energy\s+\(RHF\) : (-?\d+\.\d+)",
        "UHF-CCSD": r">>> Running UHF-CCSD <<<.*?CCSD correlation energy\s+\(UHF\) : (-?\d+\.\d+).*?CCSD total energy\s+\(UHF\) : (-?\d+\.\d+)",
        "RHF-CCSD(T)": r">>> Running RHF-CCSD <<<.*?CCSD\(T\) correlation energy \(RHF\) : (-?\d+\.\d+).*?CCSD\(T\) total energy\s+\(RHF\) : (-?\d+\.\d+)",
        "UHF-CCSD(T)": r">>> Running UHF-CCSD <<<.*?CCSD\(T\) correlation energy \(UHF\) : (-?\d+\.\d+).*?CCSD\(T\) total energy\s+\(UHF\) : (-?\d+\.\d+)",
    }

    for method, pattern in patterns.items():
        match = re.search(pattern, content, re.DOTALL)
        if match:
            if method in ("RHF", "UHF"):
                results[method] = float(match.group(1))
            else:
                result_method = cast(MethodResultType, results[method])
                result_method["corr"] = float(match.group(1))
                result_method["tot"] = float(match.group(2))
    # print(results)
    return results


def are_floats_equal(a: float, b: float, epsilon: float = 1.0001e-6) -> bool:
    # print(f"{a=}, {b=}, {abs(a - b)=}, {epsilon=}")
    return abs(a - b) < epsilon


def _format_cebe(cebe: float) -> float:
    return float(f"{cebe:.6f}")


def compare_delta_to_sp_dicts(
    delta_dict: DeltaEnergyType, sp_dict: StateResults
) -> bool:
    delta_hf_sp = HA_TO_EV * (sp_dict["UHF"] - sp_dict["RHF"])
    if not are_floats_equal(_format_cebe(delta_hf_sp), delta_dict["UHF"]):
        return False
    if "MP2" not in delta_dict:
        return True
    delta_mp2_sp = HA_TO_EV * (sp_dict["UHF-MP2"]["tot"] - sp_dict["RHF-MP2"]["tot"])
    if not are_floats_equal(_format_cebe(delta_mp2_sp), delta_dict["MP2"]):
        return False

    if "CCSD" not in delta_dict:
        return True
    delta_ccsd_sp = HA_TO_EV * (sp_dict["UHF-CCSD"]["tot"] - sp_dict["RHF-CCSD"]["tot"])
    if not are_floats_equal(_format_cebe(delta_ccsd_sp), delta_dict["CCSD"]):
        return False

    if "CCSD(T)" not in delta_dict:
        return True
    delta_ccsdt_sp = HA_TO_EV * (
        sp_dict["UHF-CCSD(T)"]["tot"] - sp_dict["RHF-CCSD(T)"]["tot"]
    )
    if not are_floats_equal(_format_cebe(delta_ccsdt_sp), delta_dict["CCSD(T)"]):
        return False
    return True


IGNORED_BASES = {"pcX-2", "pcX-3", "pcX-4", "ccX-TZ", "ccX-QZ", "ccX-5Z"}

if __name__ == "__main__":
    path = "/Users/morgunov/vv/cebe_prediction/data/calculations/mom/o/co/pcX-1/pyscf_output_mom.txt"
    sp_energies = parse_pyscf_output(path)
    delta_E: DeltaEnergyType = {
        "UHF": 541.396936,
        "MP2": 542.994313,
        "CCSD": 542.753663,
        "CCSD(T)": 542.828197,
    }
    print(compare_delta_to_sp_dicts(delta_E, sp_energies))


def get_next_col(col: str, prefix: Optional[str] = None) -> str:
    if prefix is None:
        prefix = ""
    strings = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
    if not col:
        return "A"
    if len(col) == 1:
        if col == "Z":
            return get_next_col(prefix, prefix=None) + get_next_col("", prefix=None)
        else:
            return prefix + strings[strings.index(col) + 1]
    else:
        return get_next_col(col[1:], prefix=prefix + col[0])


class ParsedData:
    """
    This object takes an excel file with parsed data and updates it with new results. The advantage of this approach is that any custom comments left in column L in CEBE_Data file are preserved.
    """

    # fmt:off
    specialBases = {"ccX-DZ","ccX-TZ","ccX-QZ","ccX-5Z","pcX-1","pcX-2","pcX-3","pcX-4",}  # fmt:on

    def __init__(
        self,
        experimental_wb: str,
        calculations_wb: str,
        algorithms: List[str],
        calculations_folder: Path,
        do_average_experimental: bool = True,
    ):
        """
        Creates a data object. Main reason for using a class is to avoid having to pass different data objects from function to funtion
        """
        self.experWB = load_workbook(experimental_wb, data_only=True)
        self.exper_wb_col = "L"
        self.readWB = load_workbook(calculations_wb)
        self.save_path = calculations_wb
        self.calc_path = calculations_folder

        self.saveWB = Workbook()

        self.molToExper: ExperDataType = {}
        self.doAverageExperimental = do_average_experimental

        self.algorithms = algorithms
        self.algoToDelta: AlgoDeltaType = {}
        self.algoToEnergy: AlgoEnergyType = {}

        self.debug = False

        self.font = Font(name="Helvetica", size=12)
        self.cols = "B C D E F G H I J K L M N O P Q R S"
        # algo -> atom -> molecule -> basis -> method -> data (str, float)

    def _parse_calculations(self) -> None:
        """
        Parse data from folders listed in self.algorithmToFolders and store it in a molToData dictionary

        molToData maps molecule (str) to basis (D/T, also str) to a dictionary:
            \{
                "atom": atom that is excited
                "error": error if any
                "UHF": result of UHF calculation
                "MP2": result of MP2 calculation
                "MP3": result of MP3 calculation
                "CCSD": result of CCSD calculation
            \}

        molToData is saved as a value of self.algoToDelta with relevant key

        P.S. This is already implemented to work with different algorithms (mom, sgm, coreless). Just use a different key in algorithmToFolders
        """
        algoToDelta = self.algoToDelta
        for algorithm in self.algorithms:
            algo_path = self.calc_path / algorithm
            atoms = [item.name for item in algo_path.glob("*/")]
            if self.debug:
                print(f"{atoms=}")
            for atom in atoms:
                molToData = algoToDelta.setdefault(algorithm, {}).setdefault(atom, {})
                if self.debug:
                    print(f"{self.algoToDelta=}, {molToData=}")
                atom_path = algo_path / atom
                molecules = [molecule.name for molecule in atom_path.glob("*/")]
                if self.debug:
                    print(f"{molecules=}")
                for molecule in molecules:
                    if self.atom_to_mols is not None and (
                        atom not in self.atom_to_mols
                        or molecule not in self.atom_to_mols[atom]
                    ):
                        continue
                    mol_path = atom_path / molecule
                    bases = [basis.name for basis in mol_path.glob("*/")]
                    if self.debug:
                        print(f"{bases=}")
                    for basis in bases:
                        basisToData = molToData.setdefault(molecule, {})
                        data_point = basisToData.setdefault(basis, {})
                        bas_path = mol_path / basis
                        result_file = bas_path / f"CEBE_{algorithm}.txt"
                        pyscf_out = bas_path / f"pyscf_output_{algorithm}.txt"

                        # Case 1. Calculation terminated normally
                        if result_file.exists() and open(result_file, "r").read() != "":
                            lines = open(result_file, "r").readlines()
                            for rline in lines:
                                line = rline.split("\n")[0]
                                if ":" in line:
                                    _line_comps = rline.split("\n")[0].split(":")
                                    if _line_comps[0] not in allowed_extra_keys:
                                        raise KeyError(
                                            f"Received unrecognized entry {_line_comps[0]}, expected one of {allowed_extra_keys}"
                                        )
                                    key: LitKeyType = cast(LitKeyType, _line_comps[0])
                                    val: str = _line_comps[1]
                                    data_point[key] = val
                                    continue
                                if "=" not in line:
                                    continue
                                rmethod, rvalue = line.split(" = ")
                                value = float(rvalue.split(" ")[0])
                                posthf_method = cast(
                                    LitKeyType, rmethod.split(" ")[1][1:-1]
                                )
                                data_point[posthf_method] = value
                            if not pyscf_out.exists():
                                raise FileNotFoundError(
                                    f"Expected {pyscf_out} to exist"
                                )
                            sp_energies = parse_pyscf_output(pyscf_out)
                            self.algoToEnergy.setdefault(algorithm, {}).setdefault(
                                atom, {}
                            ).setdefault(molecule, {})[basis] = sp_energies
                            if not compare_delta_to_sp_dicts(data_point, sp_energies):
                                if basis in IGNORED_BASES:
                                    continue
                                print(f"\n\n{molecule=} {basis=}")
                                print(f"{data_point=}")
                                print(f"{sp_energies=}")
                                raise ValueError(
                                    f"Data from pyscf output {pyscf_out} does not match data from CEBE file {result_file}"
                                )

                        # Case 2. Calculation was unsuccesful and did not terminate normally
                        else:
                            data_point["error"] = "No CEBE file"
                            err_file = bas_path / "sbatch.err"
                            if err_file.exists():
                                errFile = open(err_file, "r").readlines()
                                if errFile:
                                    data_point["detailed"] = errFile[-1].split("\n")[0]
                                else:
                                    data_point["detailed"] = "empty?"
                            else:
                                data_point["detailed"] = "empty?"
                                continue

    def _parse_experimental(self) -> None:
        """
        Parse data from experimental_wb file and create a dictionary mapping molecule name (which is used to denote file names) to experimental values
        """
        ws = self.experWB["Sheet1"]
        data = {}
        row = 2
        while True:
            if ws["A" + str(row)].value is None:
                # empty row reached = end of file
                break
            fname = ws["B" + str(row)].value
            data[fname] = ws[self.exper_wb_col + str(row)].value
            row += 1
        self.molToExper = data

    def _calculate_mp25(self) -> None:
        for atomData in self.algoToDelta.values():
            for molData in atomData.values():
                for basData in molData.values():
                    for data in basData.values():
                        if "MP2" in data and "MP3" in data:
                            if not isinstance(data["MP2"], float):
                                raise TypeError(
                                    f"Expected float for MP2, got {data['MP2']}"
                                )
                            if not isinstance(data["MP3"], float):
                                raise TypeError(
                                    f"Expected float for MP3, got {data['MP3']}"
                                )
                            data["MP2.5"] = (data["MP2"] + data["MP3"]) / 2

    def _update_calculation_wb(self) -> None:
        """
        Write the contents of self.algoToDelta from parse_results to a CEBE_Data file, preserving any comments in column L (currently not used)
        """
        readWS = self.readWB["Sheet"]
        ws = self.saveWB["Sheet"]

        r = 3
        for col, name in CALC_WB_COLS.items():
            ws[col + str(r)] = name

        for col in self.cols.split(" "):
            ws[col + str(r)].font = self.font
        r += 1
        for algorithm, atomData in self.algoToDelta.items():
            for atom, molData in atomData.items():
                for molecule, basToData in molData.items():
                    for basis, data in basToData.items():
                        ws["B" + str(r)] = molecule
                        ws["C" + str(r)] = algorithm
                        ws["D" + str(r)] = basis
                        ws["E" + str(r)] = atom
                        if molecule in self.molToExper:
                            ws["K" + str(r)] = self.molToExper[molecule]
                        if "error" in data:
                            ws["Q" + str(r)] = data["error"]
                            ws["R" + str(r)] = data["detailed"]
                        else:
                            if "UHF" in data:
                                ws["F" + str(r)] = data["UHF"]
                            if "MP2" in data:
                                ws["G" + str(r)] = data["MP2"]
                            if "MP3" in data:
                                ws["H" + str(r)] = data["MP3"]
                            if "CCSD" in data:
                                ws["I" + str(r)] = data["CCSD"]
                            if "CCSD(T)" in data:
                                ws["J" + str(r)] = data["CCSD(T)"]
                            if "CCSDT" in data:
                                ws["J" + str(r)] = data["CCSDT"]
                            if "Frozen orbitals" in data:
                                ws["L" + str(r)] = data["Frozen orbitals"]
                            if "Swapped orbitals" in data:
                                ws["M" + str(r)] = data["Swapped orbitals"]
                            if "T1 for RHF" in data:
                                ws["N" + str(r)] = round(float(data["T1 for RHF"]), 4)
                            if "T1 for UHF(a)" in data:
                                ws["O" + str(r)] = round(
                                    float(data["T1 for UHF(a)"]), 4
                                )
                            if "T1 for UHF(b)" in data:
                                ws["P" + str(r)] = round(
                                    float(data["T1 for UHF(b)"]), 4
                                )
                        ws["S" + str(r)] = readWS["S" + str(r)].value
                        for col in self.cols.split(" "):
                            ws[col + str(r)].font = self.font

                        r += 1

        self.saveWB.save(self.save_path)

    def extract_molecules(self, mols: Set[str], save_path: str) -> None:
        """ """
        new_wb = Workbook()
        new_ws = new_wb["Sheet"]

        wb = load_workbook(self.save_path)
        ws = wb["Sheet"]

        row = 3
        for col, name in CALC_WB_COLS.items():
            new_ws[col + str(row)] = name

        for col in self.cols.split(" "):
            new_ws[col + str(3)].font = self.font
        row, new_row = 4, 4
        while True:
            if (mol := ws["B" + str(row)].value) is None:
                break
            if mol in mols:
                new_ws["B" + str(new_row)] = mol
                for col in self.cols.split():
                    new_ws[col + str(new_row)] = ws[col + str(row)].value
                for col in self.cols.split(" "):
                    new_ws[col + str(new_row)].font = self.font
                new_row += 1

            row += 1
        new_wb.save(save_path)

    def filter_data_by_molecules(
        self, algoToDelta: AlgoDeltaType, atomToMols: Dict[str, Set[str]]
    ) -> AlgoDeltaType:
        filtered_algoToData: AlgoDeltaType = {}
        for algorithm, atomData in algoToDelta.items():
            for atom, molData in atomData.items():
                if atom not in atomToMols:
                    continue
                for molecule, basData in molData.items():
                    if molecule in atomToMols[atom]:
                        filtered_algoToData.setdefault(algorithm, {}).setdefault(
                            atom, {}
                        )[molecule] = basData
                        filtered_algoToData[algorithm][atom][molecule] = basData
        return filtered_algoToData

    def filter_by_presence_of_experimental(
        self, algoToDelta: AlgoDeltaType
    ) -> AlgoDeltaType:
        filtered_algoToData: AlgoDeltaType = {}
        for algorithm, atomData in algoToDelta.items():
            for atom, molData in atomData.items():
                for molecule, basData in molData.items():
                    if molecule in self.molToExper:
                        filtered_algoToData.setdefault(algorithm, {}).setdefault(
                            atom, {}
                        )[molecule] = basData
        return filtered_algoToData

    def calculate_errors(self, algoToDelta: AlgoDeltaType) -> None:
        valid_keys = {"UHF", "MP2", "MP2.5", "MP3", "CCSD", "CCSD(T)"}
        algoToError: AlgoErrorType = {}
        for algorithm, atomData in algoToDelta.items():
            for atom, molData in atomData.items():
                for molecule, basData in molData.items():
                    if molecule not in self.molToExper:
                        continue
                    for bas, methodData in basData.items():
                        for key, value in methodData.items():
                            if key not in valid_keys:
                                continue
                            if not isinstance(value, float):
                                raise TypeError(f"Expected float, got {value}")
                            error = value - self.molToExper[molecule]
                            if self.debug:
                                if error > 1 and bas in self.specialBases:
                                    print(
                                        f"Large error for {algorithm} {atom} {molecule} {bas} {key}: {error}"
                                    )
                            algoToError.setdefault(algorithm, {}).setdefault(
                                atom, {}
                            ).setdefault(molecule, {}).setdefault(bas, {})[key] = error
        self.algoToError = algoToError

    def calculate_series_statistics(self) -> None:
        if not hasattr(self, "algoToError"):
            raise AttributeError("You must call calculate_errors first")
        algoToAtomErrors: AlgoAtomErrorsType = {}
        for algorithm, atomData in self.algoToError.items():
            for atom, molData in atomData.items():
                for basData in molData.values():
                    for bas, methodData in basData.items():
                        for method, value in methodData.items():
                            algoToAtomErrors.setdefault(algorithm, {}).setdefault(
                                atom, {}
                            ).setdefault(bas, {}).setdefault(method, []).append(value)
        algoToAtomStats: AlgoAtomStatsType = {}
        for algorithm, atomErrData in algoToAtomErrors.items():
            for atom, basErrData in atomErrData.items():
                for bas, methodErrData in basErrData.items():
                    for method, error_list in methodErrData.items():
                        errors = np.array(error_list)
                        abs_errs = np.abs(errors)
                        algoToAtomStats.setdefault(algorithm, {}).setdefault(
                            atom, {}
                        ).setdefault(bas, {})[method] = {
                            "MSE": np.mean(errors),
                            "MAE": np.mean(abs_errs),
                            "MedAE": np.median(abs_errs),
                            "MaxAE": np.max(abs_errs),
                            "STD(AE)": np.std(abs_errs),
                            "n": len(errors),
                            "errors": errors,
                            "abs_errors": abs_errs,
                        }
        self.algoToAtomStats = algoToAtomStats

    def calculate_overall_statistics(self) -> None:
        if not hasattr(self, "algoToError"):
            raise AttributeError("You must call calculate_errors first")
        algoToErrors: AlgoErrorsType = {}
        for algorithm, atomData in self.algoToError.items():
            for molData in atomData.values():
                for basData in molData.values():
                    for bas, methodData in basData.items():
                        for method, value in methodData.items():
                            algoToErrors.setdefault(algorithm, {}).setdefault(
                                bas, {}
                            ).setdefault(method, []).append(value)
        algoToStats: AlgoStatsType = {}
        for algorithm, basErrData in algoToErrors.items():
            for bas, methodErrData in basErrData.items():
                for method, error_list in methodErrData.items():
                    errors = np.array(error_list)
                    abs_errs = np.abs(errors)
                    algoToStats.setdefault(algorithm, {}).setdefault(bas, {})[
                        method
                    ] = {
                        "MSE": np.mean(errors),
                        "MAE": np.mean(abs_errs),
                        "MedAE": np.median(abs_errs),
                        "MaxAE": np.max(abs_errs),
                        "STD(AE)": np.std(abs_errs),
                        "n": len(errors),
                        "errors": errors,
                        "abs_errors": abs_errs,
                    }
        self.algoToStats = algoToStats

    def process(
        self, atom_to_mols: Optional[Dict[str, Set[str]]] = None, save: bool = True
    ) -> None:
        self.atom_to_mols = atom_to_mols
        self._parse_experimental()
        self._parse_calculations()
        self._calculate_mp25()
        if save:
            self._update_calculation_wb()


if __name__ == "__main__":
    # perform_parsing()
    pass
