# import my_constants as constants
import constants
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Font
import numpy as np
import scipy
from analysis.regression.submodules import graph 
import sys
sys.path.append('../../core_excitations') # this allows us to import files from parent directories
import PRIVATE
import tools
import analysis.regression.old_regress
import analysis.parser.parse
import pprint

class EvaluateObj:
    """Object which parses necessary data to compare predictions with experimental data
    """
    def __init__(self, atomToMols):
        """Initiation

        Args:
            source (dict): atoms to mols from constants
        """
        self.atomToMols = atomToMols
        self.parserObj = analysis.parser.parse.perform_parsing()
        self.molToExper = self.parserObj.molToExper
        self.algorithmToData = self.parserObj.methodToData # method -> molecule -> basis -> data
        self.atomToObj = {}
        # self.molToExper = {}
        self.atomToExpers = {}
        self.bases = ('D', 'T', 'Q', '5')
        self.basisToMethods = {
            'D': ('UHF', 'MP2', 'MP2.5', 'MP3', 'CCSD', 'CCSD(T)'),
            'T': ('UHF', 'MP2', 'MP2.5', 'MP3', 'CCSD', 'CCSD(T)'),
            'Q': ('UHF', 'MP2', 'CCSD', 'CCSD(T)'),
            '5': ('UHF', 'MP2'),
        }
        self.methods = {'UHF', 'MP2', 'MP3', 'CCSD', 'CCSD(T)'}
        self.path = tools.join_path(PRIVATE.BASE_PATH + ["analysis", "regression"], endWithSlash=True)
        self.rounder = tools.rounder

    def preprocess(self):
        self.molToAtom = {}
        self.molWithExpToAtom = {}
        for atom, mols in self.atomToMols.items():
            for mol in mols.split():
                assert mol not in self.molToAtom, f'{mol} is a duplicate name'
                if mol in self.molToExper:
                    self.molWithExpToAtom[mol] = atom
                self.molToAtom[mol] = atom

    def compare_experimental(self, atoms=set(), molfilter = None):
        """Creates a dictionary mapping basis to method to deviations (predict-exper)

        Args:
            atoms (tuple): a tuple of atoms which we include in data analysis

        Returns:
            dict: basisToDeviations
        """
        basisToDev = {}
        if molfilter is None: molfilter = self.molWithExpToAtom
        for algorithm, molToData in self.algorithmToData.items():
            for molecule, basToData in molToData.items():
                if molecule not in molfilter: continue
                if self.molToAtom[molecule] not in atoms: continue
                for basis, data in basToData.items():
                    if basis not in self.basisToMethods: continue # ignore small basis 4-31G\STO-3G
                    for method in self.basisToMethods[basis]:
                        if method not in data: print(molecule, basis, method)
                        predict = data[method]
                        exper = self.molToExper[molecule]
                        dev = predict - exper
                        basisToDev.setdefault(basis, {}).setdefault(method, []).append(dev)
        return basisToDev

    def analyze_deviations(self, data):
        """Data is an output from self.compare_experimental()

        Args:
            data (dict): basisToDev

        Returns:
            basisToStats: dictionary mapping basis/method to mean and std
        """
        basisToStats = {}
        for basis, methodToDevs in data.items():
            if basis not in basisToStats:
                basisToStats[basis] = {}
            for method, devs in methodToDevs.items():
                absDevs = [abs(dev) for dev in devs]
                basisToStats[basis][method] = {'meanAE': np.mean(absDevs), 'medAE': np.median(absDevs),
                        'maxAE': max(absDevs), 'std': np.std(absDevs), 
                        'sample-size': len(devs), 
                        'mse': np.mean(devs)} #MSE - mean signed error
        return basisToStats

    def generate_atom_combinations(self, curCombo=None, atoms=None):
        """A generator creating atom combinations

        Args:
            curCombo (tuple, optional): a combination that is later yielded. Defaults to None.
            atoms (list, optional): a list of atoms that can be used. Defaults to None.

        Yields:
            tuple: curCombo
        """
        # print(f"calling with {curCombo} and {atoms}")
        if curCombo is None:
            curCombo = tuple()
        if atoms is None:
            atoms = list(self.atomToMols.keys())
        if atoms == []:
            yield curCombo
        else:
            # print(atoms[0])
            yield from self.generate_atom_combinations(curCombo + (atoms[-1],), atoms[:-1])
            yield from self.generate_atom_combinations(curCombo, atoms[:-1])

    def summary_of_deviations(self, atomCombos):
        """Writes summary of deviations to a markdown file
        """
        round2 = self.rounder(2)
        _fn = lambda x: round2(x)
        _fo = lambda x: f"{round2(x)}"
        _me = lambda x: f"$\Delta${x}" if x != 'UHF' else f"$\Delta$HF"
        f = open(tools.append_path(self.path, ["exports", "deviations_summary.md"]), 'w')
        # atomCombos = [combo for combo in self.generate_atom_combinations()]
        # atomCombos.sort(key= lambda a : len(a))
        atomsBodies = []
        header = ['Basis', 'Method', 'MSE', 'MAE', 'MedAE', 'MaxAE', 'STD', 'Sample Size']
        for atoms in atomCombos:
            if not atoms: continue
            f.write(f'### ')
            for atom in atoms: 
                f.write(f'{atom}')
            f.write('\n')
            
            data = self.compare_experimental(atoms)
            basisToStats = self.analyze_deviations(data)
            f.write(f"\n| Basis | Method | MSE | MaxAE | MedAE | MAE | STD(AE) | Sample Size |\n")
            f.write("|-|-|-|-|-|-|-|-|\n")
            body = []
            for basis in self.bases:
                if basis in basisToStats:
                    for method in self.basisToMethods[basis]:
                        if method in basisToStats[basis]:
                            stats = basisToStats[basis][method]
                            f.write(f"|{basis}|{method}|{_fn(stats['mse'])}|{_fn(stats['maxAE'])}|{_fn(stats['medAE'])}|{_fn(stats['meanAE'])}|{_fn(stats['std'])}|{stats['sample-size']}|\n")
                            row = [basis, _me(method), _fo(stats['mse']), _fo(stats['meanAE']), _fo(stats['medAE']), _fo(stats['maxAE']), _fo(stats['std']), f"{stats['sample-size']}"]
                            body.append(row)
            atomsBodies.append((atoms, body))
        return header, atomsBodies

    def export_deviations_to_excel(self):
        """Exports deviations to excel file
        """
        wb = Workbook()
        font = Font(name = 'Helvetica', size=12)
        # self.atomToExpers[atom][mol][basis]['exper'] = exper 
        for atom, molToData in self.atomToExpers.items():
            wb.create_sheet(atom)
            ws = wb[atom]
            row = 2
            cols = 'B C D E F G H I J K L M N O'
            ws['B'+str(row)] = 'Atom'
            ws['C'+str(row)] = 'Molecule'
            ws['D'+str(row)] = 'Basis'
            ws['E'+str(row)] = 'Experimental'
            ws['F'+str(row)] = 'UHF'
            ws['G'+str(row)] = 'dUHF-E'
            ws['H'+str(row)] = 'MP2'
            ws['I'+str(row)] = 'dMP2-E'
            ws['J'+str(row)] = 'MP3'
            ws['K'+str(row)] = 'dMP3-E'
            ws['L'+str(row)] = 'CCSD'
            ws['M'+str(row)] = 'dCCSD-E'
            ws['N'+str(row)] = 'CCSD(T)'
            ws['O'+str(row)] = 'dCCSD(T)-E'
            for col in cols.split(' '):
                ws[col+str(row)].font = font
            for mol, basToData in molToData.items():
                for bas, data in basToData.items():
                    row += 1
                    ws['B'+str(row)] = atom
                    ws['C'+str(row)] = mol
                    ws['D'+str(row)] = bas
                    ws['E'+str(row)] = data['exper']
                    ws['F'+str(row)] = data['UHF']
                    ws['G'+str(row)] = data['UHF'] - data['exper']
                    if 'MP2' in data:
                        ws['H'+str(row)] = data['MP2']
                        ws['I'+str(row)] = data['MP2'] - data['exper']
                    if 'MP3' in data:
                        ws['J'+str(row)] = data['MP3']
                        ws['K'+str(row)] = data['MP3'] - data['exper']
                    if 'CCSD' in data:
                        ws['L'+str(row)] = data['CCSD']
                        ws['M'+str(row)] = data['CCSD'] - data['exper']
                    if 'CCSD(T)' in data:
                        ws['N'+str(row)] = data['CCSD(T)']
                        ws['O'+str(row)] = data['CCSD(T)'] - data['exper']
                    for col in cols.split(' '):
                        ws[col+str(row)].font = font
        wb.save(tools.append_path(self.path, ["exports", "evaluation", "summary.xlsx"]))

    def assess_regression(self, basis, xMethod, yMethod, atomLearn, atomsEval):
        """Evaluates how good a method is on cross-analysis

        Args:
            basis (str): zeta size of basis set
            xMethod (str): method1
            yMethod (str): method2
            atomLearn (str): atom on which regression is built
            atomsEval (str): atom on which regression is evaluated

        Returns:
            tuple: statistics
        """
        # perform t-test (Student's t-test)
        model = self.atomToObj[atomLearn].methodsToRegress[basis][xMethod][yMethod]

        predictDevs = []
        for atom in atomsEval:
            evalData = self.atomToObj[atom].parsed
            for mol, basisToData in evalData.items():
                if basis in basisToData:
                    predict = basisToData[basis][xMethod] * model['slope'] + model['intercept']
                    real = basisToData[basis][yMethod]
                    delta = predict-real
                    predictDevs.append(delta)

        # Welch's t-test
        evalDevs = []
        for mol, basisToData in self.atomToObj[atomLearn].parsed.items():
            if basis in basisToData:
                predict = basisToData[basis][xMethod] * model['slope'] + model['intercept']
                real = basisToData[basis][yMethod]
                delta = predict-real
                evalDevs.append(delta)
        
        studentT = scipy.stats.ttest_ind(predictDevs, [0]*len(predictDevs))
        welchT = scipy.stats.ttest_ind(predictDevs, evalDevs)

        # Shapiro-Wilk test for variance
        shapiroP = scipy.stats.shapiro(predictDevs)
        shapiroE = scipy.stats.shapiro(evalDevs)

        # F - test
        fTest = scipy.stats.f_oneway(predictDevs, evalDevs)

        # KS - test
        ksTest = scipy.stats.kstest(predictDevs, evalDevs)
        fname = f"{basis}-{xMethod}-{yMethod}-{atomLearn}-{'_'.join(atomsEval)}"
        return studentT, welchT, shapiroP, shapiroE, fTest, ksTest, fname
    
    def plot_model_against_data(self, basis, xMethod, yMethod, atomLearn, atomsEval):
        """Plots model evaluations

        Args:
            basis (str): zeta basis size
            xMethod (str): method1
            yMethod (str): method2
            atomLearn (str): atom on which regression is built
            atomsEval (str): atom on which regressionn is evaluated
        """
        model = self.atomToObj[atomLearn].methodsToRegress[basis][xMethod][yMethod]
        xVals, yVals, yLine = [], [],[]
        for atom in atomsEval:
            evalData = self.atomToObj[atom].parsed
            for mol, basisToData in evalData.items():
                if basis in basisToData:
                    predict = basisToData[basis][xMethod] * model['slope'] + model['intercept']
                    real = basisToData[basis][yMethod]
                    xVals.append(basisToData[basis][xMethod])
                    yVals.append(real)
                    yLine.append(predict)

        #fname, xAxis, yAxis, title, dataPairs):
        fname = f"{basis}-{xMethod}-{yMethod}-{atomLearn}-{'_'.join(atomsEval)}"
        xAxis = f"{xMethod} (eV)"
        yAxis = f"{yMethod} (ev)"
        title = f"Evaluating {xMethod} regression built on {atomLearn} against {yMethod} of {'/'.join(atomsEval)}"
        # xVals, yVals, label, mode, palette
        dataPairs = tuple()
        dataPairs += ((xVals, yVals, "Calculated", "markers", "markers"),)
        dataPairs += ((xVals, yLine, "Predicted", "lines", "red"),)
        graphObj = graph.GraphObject(fname, xAxis, yAxis, title, dataPairs)
        graphObj.plot()
                    
    def create_summary(self):
        # return studentT, welchT, shapiroP, shapiroE, fTest, ksTest, fname
        atoms = tuple(self.atomToObj.keys())
        f = open(tools.append_path(self.path, ["exports", "evaluations", "md", "summary.md"]), "w")
        f.write("## Summary\n")
        f.write("| Basis | xMethod | yMethod | atomLearn | atomEval | student T | welch T | Shapiro P | Shapiro E | F-Test | KS Test |\n")
        f.write("| - | - | - | - | - | - | - | - | - | - | - |\n")
        for basis in self.bases:
            if basis not in {'D',}:
                for method in self.methods:
                    if method not in {'ccsd', 'uhf'}:
                        for atom in atoms:
                            # for atomEvals in self.generate_atom_combinations():
                            #     if atomEvals:
                            for atomEvals in atoms:
                                self.plot_model_against_data(basis, method, 'ccsd', atom, (atomEvals, ))
                                studentT, welchT, shapiroP, shapiroE, fTest, ksTest, fname = self.assess_regression(basis, method, 'ccsd', atom, (atomEvals,))
                                # fname = f"{basis}-{xMethod}-{yMethod}-{atomLearn}-{'_'.join(atomsEval)}"
                                basis, xMethod, yMethod, atomLearn, atomEval = fname.split('-')
                                if shapiroP.pvalue < 0.05 or shapiroE.pvalue < 0.05:
                                    fTestInd = 'N/A'
                                else:
                                    fTestInd = fTest.pvalue > 0.05
                                f.write(f"| {basis} | {xMethod} | {yMethod} | {atomLearn} | {atomEval} | {studentT.pvalue > 0.05} | {welchT.pvalue > 0.05} | {shapiroP.pvalue > 0.05} | {shapiroE.pvalue > 0.05} | {fTestInd} | {ksTest.pvalue > 0.05} |\n")

    def main(self):
        self.preprocess()

def export():
    eobj = EvaluateObj(constants.ATOM_TO_MOLS)
    eobj.main()
    return eobj

    
if __name__ == "__main__":
    eobj = EvaluateObj(constants.ATOM_TO_MOLS)
    eobj.main()

