from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font
import os
import sys

sys.path.append(
    "../core_excitations"
)  # this allows us to import files from parent directories
import PRIVATE
import tools
import numpy as np


class ParserObject:
    """
    This object takes an excel file with parsed data and updates it with new results. The advantage of this approach is that any custom comments left in column L in CEBE_Data file are preserved.
    """

    def __init__(self, path: str, file: str, experFile: str, methodToFolders):
        """
        Creates a data object. Main reason for using a class is to avoid having to pass different data objects from function to funtion

        Args:
            path (str): path to parent folder (where results of calculations are also stored)
            file (str): path to file with parsed data

        Note that path and file are combined immediately after receiving. In sum, there isn't much reason to store them as separate variables, but I imagine it could be handy at some point in the future
        """
        self.BASE = path
        self.SUBMODULE = self.BASE + ["analysis", "parser"]
        self.CALCULATIONS = self.BASE + ["calculations"]
        self.readWB = load_workbook(
            tools.join_path(
                self.SUBMODULE
                + [
                    file,
                ]
            )
        )
        self.saveWB = Workbook()
        # self.parentPath = path
        self.readFile = file
        self.experFile = experFile
        self.methodToData = {}
        self.methodToFolders = methodToFolders
        self.molToExper = {}
        self.molPaths = set()
        self.molToData = {}
        self.DO_AVERAGING = True

    def parse_folders(self):
        """
        Parse data from folders listed in self.methodToFolders and store it in a molToData dictionary

        molToData maps molecule (str) to basis (D/T, also str) to a dictionary:
            \{
                "atom": atom that is excited
                "error": error if any
                "UHF": result of UHF calculation
                "MP2": result of MP2 calculation
                "MP3": result of MP3 calculation
                "CCSD": result of CCSD calculation
            \}

        molToData is saved as a value of self.methodToData with relevant key

        P.S. This is already implemented to work with different methods (mom, sgm, coreless). Just use a different key in methodToFolders
        """
        for method, folders in self.methodToFolders.items():
            for folder in folders:
                mefolder = tools.join_path([method, folder])
                if mefolder not in self.methodToData:
                    self.methodToData[mefolder] = {}
                molToData = self.methodToData[mefolder]
                molPath = tools.join_path(
                    self.CALCULATIONS
                    + [
                        mefolder,
                    ],
                    endWithSlash=True,
                )
                molecules = [
                    molecule
                    for molecule in os.listdir(molPath)
                    if os.path.isdir(
                        tools.append_path(
                            molPath,
                            [
                                molecule,
                            ],
                            endWithSlash=True,
                        )
                    )
                ]
                for molecule in molecules:
                    self.molPaths.add(folder + os.sep + molecule)
                    if molecule not in molToData:
                        molToData[molecule] = {}

                    bases = [
                        basis
                        for basis in os.listdir(
                            tools.append_path(
                                molPath,
                                [
                                    molecule,
                                ],
                                endWithSlash=True,
                            )
                        )
                        if os.path.isdir(
                            tools.append_path(
                                molPath, [molecule, basis], endWithSlash=True
                            )
                        )
                    ]
                    for basis in bases:
                        if basis not in molToData[molecule]:
                            molToData[molecule][basis] = {}

                        files = os.listdir(
                            tools.append_path(
                                molPath, [molecule, basis], endWithSlash=True
                            )
                        )
                        submitScript = open(
                            tools.append_path(molPath, [molecule, basis, "submit.sh"]),
                            "r",
                        )
                        atom = submitScript.readlines()[-1].split(" ")[2]
                        molToData[molecule][basis]["atom"] = atom
                        if f"CEBE_{method}.txt" not in files:
                            molToData[molecule][basis]["error"] = "No CEBE file"
                            errFile = open(
                                tools.append_path(
                                    molPath, [molecule, basis, "sbatch.err"]
                                ),
                                "r",
                            ).readlines()
                            if errFile:
                                molToData[molecule][basis]["detailed"] = errFile[
                                    -1
                                ].split("\n")[0]
                            else:
                                molToData[molecule][basis]["detailed"] = "empty?"
                        else:
                            with open(
                                tools.append_path(
                                    molPath, [molecule, basis, f"CEBE_{method}.txt"]
                                ),
                                "r",
                            ) as f:
                                lines = f.readlines()
                                # print(molPath, molecule, basis, lines )
                            for rline in lines:
                                line = rline.split("\n")[0]
                                if ":" in line:
                                    key, val = rline.split("\n")[0].split(":")
                                    molToData[molecule][basis][key] = val
                                    continue
                                if "=" not in line:
                                    continue
                                rmethod, rvalue = line.split(" = ")
                                value = float(rvalue.split(" ")[0])
                                posthf_method = rmethod.split(" ")[1][1:-1]
                                molToData[molecule][basis][posthf_method] = value

    def _parse_experimental(self):
        """Parse data from experimental.xlsx file and create a dictionary mapping molecule name to experimental values"""
        source = tools.join_path(
            self.SUBMODULE
            + [
                self.experFile,
            ]
        )
        wb = load_workbook(source)
        ws = wb["Sheet1"]
        data = {}
        row = 2
        while True:
            if all([ws[col + str(row)].value is None for col in ("A", "B", "C")]):
                break
            if all([ws[col + str(row)].value is not None for col in ("A", "B", "C")]):
                if self.DO_AVERAGING:
                    exprs = [ws["B" + str(row)].value]
                    col = "C"
                    while True:
                        col = tools.get_next_col(col)
                        if (val := ws[col + str(row)].value) is None:
                            break
                        exprs.append(val)
                    assert all([type(expr) in {float, int} for expr in exprs])
                    expr = np.mean(exprs)

                else:
                    expr = ws["B" + str(row)].value
                fname = ws["C" + str(row)].value
                data[fname] = expr
            row += 1
        self.molToExper = data

    def write_results(self):
        """
        Write the contents of self.methodToData from parse_results to a CEBE_Data file, preserving any comments in column L (currently not used)
        """
        readWS = self.readWB["Sheet"]
        font = Font(name="Helvetica", size=12)
        ws = self.saveWB["Sheet"]

        cols = "B C D E F G H I J K L M N O P Q R S"
        r = 3
        ws["B" + str(r)] = "Molecule"
        ws["C" + str(r)] = "Method"
        ws["D" + str(r)] = "Basis"
        ws["E" + str(r)] = "Atom"
        ws["F" + str(r)] = "UHF"
        ws["G" + str(r)] = "MP2"
        ws["H" + str(r)] = "MP3"
        ws["I" + str(r)] = "CCSD"
        ws["J" + str(r)] = "CCSD(T)"
        ws["K" + str(r)] = "Exper."
        ws["L" + str(r)] = "Frozen"
        ws["M" + str(r)] = "Swapped"
        ws["N" + str(r)] = "T1 for RHF"
        ws["O" + str(r)] = "T1 for UHF(a)"
        ws["P" + str(r)] = "T1 for UHF(b)"
        ws["Q" + str(r)] = "Error"
        ws["R" + str(r)] = "Comments"
        ws["S" + str(r)] = "Custom Notes"

        for col in cols.split(" "):
            ws[col + str(r)].font = font
        r += 1
        for method, molToData in self.methodToData.items():
            for molecule, basToData in molToData.items():
                for basis, data in basToData.items():
                    ws["B" + str(r)] = molecule
                    ws["C" + str(r)] = method
                    ws["D" + str(r)] = basis
                    ws["E" + str(r)] = data["atom"]
                    if molecule in self.molToExper:
                        ws["K" + str(r)] = self.molToExper[molecule]
                    if "error" in data:
                        ws["Q" + str(r)] = data["error"]
                        ws["R" + str(r)] = data["detailed"]
                    else:
                        if "UHF" in data:
                            ws["F" + str(r)] = data["UHF"]
                        if "MP2" in data:
                            ws["G" + str(r)] = data["MP2"]
                        if "MP3" in data:
                            ws["H" + str(r)] = data["MP3"]
                        if "CCSD" in data:
                            ws["I" + str(r)] = data["CCSD"]
                        if "CCSD(T)" in data:
                            ws["J" + str(r)] = data["CCSD(T)"]
                        if "CCSDT" in data:
                            ws["J" + str(r)] = data["CCSDT"]
                        if "Frozen orbitals" in data:
                            ws["L" + str(r)] = data["Frozen orbitals"]
                        if "Swapped orbitals" in data:
                            ws["M" + str(r)] = data["Swapped orbitals"]
                        if "T1 for RHF" in data:
                            ws["N" + str(r)] = round(float(data["T1 for RHF"]), 4)
                        if "T1 for UHF(a)" in data:
                            ws["O" + str(r)] = round(float(data["T1 for UHF(a)"]), 4)
                        if "T1 for UHF(b)" in data:
                            ws["P" + str(r)] = round(float(data["T1 for UHF(b)"]), 4)
                    ws["S" + str(r)] = readWS["S" + str(r)].value
                    for col in cols.split(" "):
                        ws[col + str(r)].font = font

                    r += 1

        self.saveWB.save(
            tools.join_path(
                self.SUBMODULE
                + [
                    self.readFile,
                ]
            )
        )

    def create_mp25(self):
        for folder, molToData in self.methodToData.items():
            for molecule, basToData in molToData.items():
                for basis, data in basToData.items():
                    if "MP2" in data and "MP3" in data:
                        data["MP2.5"] = (data["MP2"] + data["MP3"]) / 2

    def create_mol_objects(self):
        for molPath in self.molPaths:
            molecule = (
                tools.join_path(self.CALCULATIONS + ["mols", "exp"], endWithSlash=True)
                + molPath
                + ".xyz"
            )
            molname = molPath.split(os.sep)[1]
            # mol = gto.M(atom=molecule)
            # mol.build()
            # self.molToData.setdefault(molname, {})['nelec'] = mol.nelec

    def wrapper(self, fillMol=False):
        self.parse_folders()
        self._parse_experimental()
        self.write_results()
        if fillMol:
            self.create_mol_objects()
        self.create_mp25()


def perform_parsing(fillMol: bool = False):
    PATH = PRIVATE.BASE_PATH
    WB = "CEBE_Data.xlsx"
    EXPER = "experimental.xlsx"
    METHOD_TO_FOLDERS = {"mom": ["f", "c", "o", "n", "cl", "s"]}
    resObj = ParserObject(
        path=PATH, file=WB, experFile=EXPER, methodToFolders=METHOD_TO_FOLDERS
    )
    resObj.wrapper(fillMol)
    return resObj


if __name__ == "__main__":
    perform_parsing()
