from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Font
import sys
sys.path.append('../core_excitations') # this allows us to import files from parent directories
import PRIVATE
import tools

class Data:
    """
        This data object allows to extract results of specific molecules from large CEBE_Data file 

        Handy when you want to compare results of specific molecules
    """

    def __init__(self, fname, path):
        """
        Args:
            fname (str): path to CEBE_Data.xlsx file
        """
        self.fname = tools.append_path(path, [fname,])
        self.BASE = path

    def extract_molecules(self, output, mols):
        """
        Args:
            output (str): file name to which we save the results
            mols (str): a string with molecules separated by spaces which results we want to see in the file
        """
        molSet = set(mols.split(' '))
        new_wb = Workbook()
        new_ws = new_wb['Sheet']
        font = Font(name = 'Helvetica', size=12)
        cols = 'B C D E F G H I J K L M N O P Q R S'

        wb = load_workbook(self.fname)
        ws = wb['Sheet']
        
        row = 3
        new_ws['B'+str(row)] = 'Molecule'
        new_ws['C'+str(row)] = 'Method'
        new_ws['D'+str(row)] = 'Basis'
        new_ws['E'+str(row)] = 'Atom'
        new_ws['F'+str(row)] = 'UHF'
        new_ws['G'+str(row)] = 'MP2'
        new_ws['H'+str(row)] = 'MP3'
        new_ws['I'+str(row)] = 'CCSD'
        new_ws['J'+str(row)] = 'CCSD(T)'
        new_ws['K'+str(row)] = 'Exper.'
        new_ws['L'+str(row)] = 'Frozen'
        new_ws['M'+str(row)] = 'Swapped'
        new_ws['N'+str(row)] = 'T1 for RHF'
        new_ws['O'+str(row)] = 'T1 for UHF(a)'
        new_ws['P'+str(row)] = 'T1 for UHF(b)'
        new_ws['Q'+str(row)] = 'Error'
        new_ws['R'+str(row)] = 'Comments'
        new_ws['S'+str(row)] = 'Custom Notes'
        for col in cols.split(' '):
                new_ws[col+str(3)].font = font
        row, new_row = 4, 4
        while True:
            if (mol:=ws['B'+str(row)].value) is None:
                break
            if mol in molSet:
                new_ws['B'+str(new_row)] = mol
                for col in cols.split():
                    new_ws[col+str(new_row)] = ws[col+str(row)].value
                for col in cols.split(' '):
                    new_ws[col+str(new_row)].font = font
                new_row += 1

            row += 1
        new_wb.save(tools.append_path(self.BASE, ["output", f"{output}.xlsx"]))


BASE = tools.join_path(PRIVATE.BASE_PATH+["analysis", "parser"], endWithSlash=True)
d = Data('CEBE_Data.xlsx', BASE)
mols = 'co co2'
fname = 'oxides'
d.extract_molecules(fname, mols)