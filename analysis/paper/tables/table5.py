import sys, Analysis.constants as constants
sys.path.append('../../core_excitations') # this allows us to import files from parent directories
import PRIVATE
import tools
import Analysis.constants as constants
import analysis.modules.table 
import analysis.regression.evaluate
from openpyxl import load_workbook
from pprint import pprint

class Tables:

    def __init__(self):
        self.latT = analysis.modules.table.LaTeXTable()
        self.evalObj = analysis.regression.evaluate.export()
        self.molToExper = self.evalObj.molToExper
        # self.atomToMols = constants.ATOM_TO_MOLS
        self.molToAtom = self.evalObj.molWithExpToAtom
        self.molToLatex = constants.FNAME_TO_MOLS
        # self.positioning = 'l l l l l l'
        self.atomToMols = {}
        for mol, atom in self.molToAtom.items():
            self.atomToMols.setdefault(atom, set()).add(mol)

        self.rounder = tools.rounder

    def parse_geometries(self):
        methodToAtomToMols = {}
        methodToBasisToMols = {}
        path = PRIVATE.BASE_PATH + ['geom_optimization', 'geometriesDB.xlsx']
        wb = load_workbook(tools.join_path(path))
        ws = wb['Sheet1']
        row = 2
        while True:
            if ws['B'+str(row)].value is None and ws['B'+str(row+1)].value is None:
                break
            if ws['B'+str(row)].value is None: 
                row += 1
                continue
            bigAtom = ws['B'+str(row)].value.upper()
            smAtom = ws['B'+str(row)].value
            atom = bigAtom[0] + smAtom[1:]
            if atom not in self.atomToMols:
                row += 1
                continue
            molecule = ws['C'+str(row)].value
            if molecule not in self.atomToMols[atom]: 
                row += 1
                continue
            method = ws['D'+str(row)].value
            basis = ws['E'+str(row)].value if ws['E'+str(row)].value is not None else 'exp'
            methodToAtomToMols.setdefault(method, {}).setdefault(atom, {}).setdefault(basis, []).append(molecule)
            methodToBasisToMols.setdefault(method, {}).setdefault(basis, set()).add(self.molToLatex[molecule]['formula'])
            row += 1
        self.methodToAtomToBasisToMols = methodToAtomToMols
        self.methodToBasisToMols = methodToBasisToMols

    def experimental_geometries(self):
        caption = f'Molecules for which experimental geometries were available'
        label = f'geom-exps'
        header = ['C-series', 'N-Series', 'O-series', 'F-series']
        positioning = 'p{0.22\linewidth} | p{0.22\linewidth} | p{0.22\linewidth} | p{0.22\linewidth}'
        path = tools.join_path(PRIVATE.BASE_PATH+['analysis', 'paper', 'output', 'tables', 'geometries', f'exps'])
        body = []
        for j, ratom in enumerate(header):
            atom = ratom[0]
            for i, molecule in enumerate(sorted(self.methodToAtomToBasisToMols['exp'][atom]['exp'])):
                cell = "\ch{%s}" % self.molToLatex[molecule]['latex']
                if i >= len(body):
                    body.append([])
                body[i].append(cell)
            
            cell = f"({len(self.methodToAtomToBasisToMols['exp'][atom]['exp'])} molecule)"
            i += 1
            if i >= len(body):
                body.append([])
            body[i].append(cell)
            i+=1
            if j > 0:
                while i < len(body):
                    body[i].append(' ')
                    i += 1
        self.latT.export_table(caption, label, positioning, header, body, path)

    def mp2full_geometries(self):
        method = 'MP2(Full)'
        caption = f'Molecules for which MP2(Full) geometries were taken'
        label = f'geom-mp2full'
        header = ['C-series', 'N-Series', 'O-series']
        positioning = 'p{0.30\linewidth} | p{0.30\linewidth} | p{0.30\linewidth}'
        path = tools.join_path(PRIVATE.BASE_PATH+['analysis', 'paper', 'output', 'tables', 'geometries', f'mp2full'])
        body = []
        prevMax = 0
        for j, ratom in enumerate(header):
            atom = ratom[0]
            i = 0
            if atom in self.methodToAtomToBasisToMols[method]:
                for basis in self.methodToAtomToBasisToMols[method][atom]:
                    if i >= len(body):
                        body.append([])
                    body[i].append("\\textit{%s}" % basis)
                    i+=1
                    for molecule in sorted(self.methodToAtomToBasisToMols[method][atom][basis]):
                        cell = "\ch{%s}" % self.molToLatex[molecule]['latex']
                        if i >= len(body):
                            body.append([])
                        body[i].append(cell)
                        i+=1
            
                    cell = f"({len(self.methodToAtomToBasisToMols[method][atom][basis])} molecule)"
                    if i >= len(body):
                        body.append([])
                    body[i].append(cell)
                    i+=1

            if j > 0:
                if i < len(body):
                    while i < len(body):
                        body[i].append(' ')
                        i += 1
                else:
                    i = prevMax
                    while i < len(body):
                        body[i].insert(0, ' ')
                        i += 1
            prevMax = i
        self.latT.export_table(caption, label, positioning, header, body, path)

    def rimp2_geometries(self):
        method = 'RI-MP2'
        caption = f'Molecules for which RI-MP2 geometries were calculated with ORCA'
        label = f'geom-rimp2'
        header = ['N-series', 'O-Series', 'F-series']
        positioning = 'p{0.30\linewidth} | p{0.30\linewidth} | p{0.30\linewidth}'
        path = tools.join_path(PRIVATE.BASE_PATH+['analysis', 'paper', 'output', 'tables', 'geometries', f'rimp2'])
        body = []
        prevMax = 0
        for j, ratom in enumerate(header):
            atom = ratom[0]
            i = 0
            if atom in self.methodToAtomToBasisToMols[method]:
                for basis in self.methodToAtomToBasisToMols[method][atom]:
                    if i >= len(body):
                        body.append([])
                    body[i].append("\\textit{%s}" % basis)
                    i+=1
                    for molecule in sorted(self.methodToAtomToBasisToMols[method][atom][basis]):
                        cell = "\ch{%s}" % self.molToLatex[molecule]['latex']
                        if i >= len(body):
                            body.append([])
                        body[i].append(cell)
                        i+=1
            
                    cell = f"({len(self.methodToAtomToBasisToMols[method][atom][basis])} molecule)"
                    if i >= len(body):
                        body.append([])
                    body[i].append(cell)
                    i+=1

            if j > 0:
                if i < len(body):
                    while i < len(body):
                        body[i].append(' ')
                        i += 1
                else:
                    i = prevMax
                    while i < len(body):
                        body[i].insert(0, ' ')
                        i += 1
            prevMax = i
        self.latT.export_table(caption, label, positioning, header, body, path)

    def summary(self):
        mols = list(sorted(self.methodToBasisToMols['exp']['exp']))
        self.methodToBasisToMols.setdefault('exp1', {})['exp1'] = mols[:len(mols)//2]
        self.methodToBasisToMols.setdefault('exp2', {})['exp2'] = mols[len(mols)//2:]
        caption = f'Molecules for which experimental geometries were available'
        label = f'geom-summary'
        header = ['exp1', 'exp2', 'MP2(Full)', 'RI-MP2']
        positioning = 'p{0.30\linewidth} | p{0.30\linewidth} | p{0.30\linewidth} | p{0.30\linewidth}'
        path = tools.join_path(PRIVATE.BASE_PATH+['analysis', 'paper', 'output', 'tables', 'geometries', f'summary'])
        body = []
        for j, method in enumerate(header):   
            i = 0     
            for basis in self.methodToBasisToMols[method]:
                if basis not in {'exp', 'exp1', 'exp2'}:
                    if i >= len(body):
                        body.append([])
                    body[i].append("\\textit{%s}" % basis)
                    i+=1
                for molecule in sorted(self.methodToBasisToMols[method][basis]):
                    cell = "\ch{%s}" % molecule
                    if i >= len(body):
                        body.append([])
                    body[i].append(cell)
                    i+=1
        
                cell = "(%s molecules)" % len(self.methodToBasisToMols[method][basis])
                if i >= len(body):
                    body.append([])
                body[i].append(cell)
                i+=1

            if j > 0:
                if i < len(body):
                    while i < len(body):
                        body[i].append(' ')
                        i += 1
        self.latT.export_table(caption, label, positioning, header, body, path)

    def all_results(self):
        for atom in 'C N O F'.split():
            for basis in 'D T Q 5'.split():
                self.series_table(atom, basis)


if __name__ == "__main__":
    tables = Tables()
    # tables.all_results()
    tables.parse_geometries()
    # tables.experimental_geometries()
    # tables.mp2full_geometries()
    # tables.rimp2_geometries()
    # pprint(tables.methodToBasisToMols)
    tables.summary()
    # pprint(tables.methodToAtomToBasisToMols['RI-MP2'])

    