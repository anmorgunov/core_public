from pathlib import Path
from typing import Callable, Dict, Iterable, List, Set, Tuple, Union, cast

from openpyxl import load_workbook

from analysis import constants
from analysis.modules import latex
from analysis.modules.extrapolation import AtomStatsType, SchemeStatsType, StatsKeyType
from analysis.modules.parsers import (
    AtomBasisStatsType,
    AtomDataType,
    BasisStatsType,
    ExperDataType,
)


class CebeTable:
    def _format(self, number: Union[int, float]) -> str:
        return f"{number:.2f}"


class SingleZetaResults(CebeTable):
    # fmt:off
    basisToHeader = {
        "5": ["Molecule", "$\Delta$HF", "$\Delta$MP2", "Experiment"],
        "Q": ["Molecule", "$\Delta$HF", "$\Delta$MP2", "$\Delta$CCSD", "$\Delta$CCSD(T)", "Experiment"],
        "T": ["Molecule", "$\Delta$HF", "$\Delta$MP2", "$\Delta$MP3", "$\Delta$CCSD", "$\Delta$CCSD(T)", "Experiment"],
        "D": ["Molecule", "$\Delta$HF", "$\Delta$MP2", "$\Delta$MP3", "$\Delta$CCSD", "$\Delta$CCSD(T)", "Experiment"],
    }
    # fmt:on

    def __init__(
        self, atomToData: AtomDataType, molToExper: ExperDataType, save_folder: Path
    ) -> None:
        self.atomToData = atomToData
        self.molToExper = molToExper
        self.save_folder = save_folder

    def _render_mol(self, mol: str) -> str:
        return constants.FNAME_TO_MOLS[mol]["latex"]

    def collect_series(self, atom: str, basis: str) -> List[List[str]]:
        body: List[List[str]] = []
        for molecule in sorted(self.atomToData[atom]):
            _data = self.atomToData[atom][molecule][basis]
            exp = self.molToExper[molecule]
            uhf, mp2 = _data["UHF"], _data["MP2"]
            row = ["\ch{%s}" % self._render_mol(molecule)]
            if basis == "5":
                energies = [uhf, mp2, exp]
            elif basis == "Q":
                ccsd, ccsdt = _data["CCSD"], _data["CCSD(T)"]
                energies = [uhf, mp2, ccsd, ccsdt, exp]
            else:
                ccsd, ccsdt = _data["CCSD"], _data["CCSD(T)"]
                mp3 = _data["MP3"]
                energies = [uhf, mp2, mp3, ccsd, ccsdt, exp]

            for e in energies:
                row.append(self._format(e))
            body.append(row)
        return body

    def series_table(self, atom: str, basis: str, save_path: str) -> None:
        body = self.collect_series(atom, basis)
        caption = f"K-Edge Ionization Energies (in eV) of {atom}-Series in cc-pV{basis}Z/cc-pCV{basis}Z"

        latex.Table.export_table(
            caption=caption,
            label=f"{atom.lower()}-{basis.lower()}z",
            positioning="l " * len(self.basisToHeader[basis]),
            headers=self.basisToHeader[basis],
            body=body,
            save_path=save_path,
        )

    def all_results(self) -> None:
        for atom in self.atomToData.keys():
            for basis in "D T Q 5".split():
                path = self.save_folder / f"{atom}-{basis}Z.tex"
                self.series_table(atom.lower(), basis, str(path))


class MethodSummary(CebeTable):
    col_names = ["Basis", "Method", "MSE", "MAE", "MedAE", "MaxAE", "STD"]
    basisToMethods = {
        "D": "UHF MP2 CCSD CCSD(T)".split(),
        "pcX-1": "UHF MP2 CCSD CCSD(T)".split(),
        "ccX-DZ": "UHF MP2 CCSD CCSD(T)".split(),
        "T": "UHF MP2 CCSD CCSD(T)".split(),
        "pcX-2": "UHF MP2 CCSD CCSD(T)".split(),
        "ccX-TZ": "UHF MP2 CCSD CCSD(T)".split(),
        "Q": "UHF MP2 CCSD CCSD(T)".split(),
        "pcX-3": "UHF MP2 CCSD CCSD(T)".split(),
        "ccX-QZ": "UHF MP2 CCSD CCSD(T)".split(),
        "5": "UHF MP2".split(),
        "pcX-4": "UHF MP2".split(),
        "ccX-5Z": "UHF MP2".split(),
    }

    def __init__(
        self,
        basisToStats: BasisStatsType,
        atomToBasisStats: AtomBasisStatsType,
        save_folder: Path,
        show_sample_size: bool = True,
        isPublication: bool = False,
    ) -> None:
        self.basisToStats = basisToStats
        self.atomToBasisStats = atomToBasisStats
        self.save_folder = save_folder
        self.show_sample_size = show_sample_size
        self.isPublication = isPublication

    def create_table_body(
        self, statsContainer: BasisStatsType, bases: List[str]
    ) -> Tuple[List[List[str]], Set[int]]:
        def _format_method(x: str) -> str:
            return f"$\Delta${x}" if x != "UHF" else "$\Delta$HF"

        body: List[List[str]] = []
        nSet: Set[int] = set()
        for basis in bases:
            if self.isPublication:
                methods = self.basisToMethods[basis]
            else:
                methods = list(statsContainer[basis].keys())
            for method in methods:
                stats = statsContainer[basis][method]
                row = [basis, _format_method(method)]
                for key in "MSE MAE MedAE MaxAE STD(AE)".split():
                    row.append(
                        self._format(cast(float, stats[cast(StatsKeyType, key)]))
                    )
                if self.show_sample_size:
                    row.append(str(stats["n"]))
                nSet.add(int(stats["n"]))
                body.append(row)
        return body, nSet

    def series_table(self, atom: str, bases: List[str], save_path: str) -> None:
        body, nSet = self.create_table_body(self.atomToBasisStats[atom], bases)

        if self.show_sample_size:
            col_names = self.col_names + ["Sample Size"]
            suffix = ""
        else:
            col_names = self.col_names
            assert len(nSet) == 1, "Sample size must be the same for all methods"
            suffix = f" ({nSet.pop()} molecules)"
        header = ["\\textbf{%s}" % col_name for col_name in col_names]
        caption = (
            f"Statistical analysis of accuracy of different methods at predicting K-Edge CEBEs (in eV) compared to experimental data for {atom.upper()}-series"
            + suffix
        )
        latex.Table.export_table(
            caption=caption,
            label=f"method-summary-{atom.lower()}",
            positioning="l " * len(header),
            headers=header,
            body=body,
            save_path=save_path,
        )

    def all_table(self, bases: List[str], save_path: str) -> None:
        body, nSet = self.create_table_body(self.basisToStats, bases)

        if self.show_sample_size:
            self.col_names.append("Sample Size")
            suffix = ""
        else:
            assert len(nSet) == 1, "Sample size must be the same for all methods"
            suffix = f" ({nSet.pop()} molecules)"
        header = ["\\textbf{%s}" % col_name for col_name in self.col_names]
        caption = (
            "Statistical analysis of accuracy of different methods at predicting K-Edge CEBEs (in eV) compared to experimental data for all data points"
            + suffix
        )
        latex.Table.export_table(
            caption=caption,
            label="method-all-summary",
            positioning="l " * len(header),
            headers=header,
            body=body,
            save_path=save_path,
        )

    def all_results(self) -> None:
        atoms = self.atomToBasisStats.keys()
        for atom in atoms:
            # if self.isPublication:
            # bases = "D T Q 5".split()
            bases = (
                "D pcX-1 ccX-DZ T pcX-2 ccX-TZ Q pcX-3 ccX-QZ 5 pcX-4 ccX-5Z".split()
            )
            # else:
            # bases = "D ccX-DZ pcX-1 T ccX-TZ pcX-2 Q ccX-QZ pcX-3 5 ccX-5Z pcX-4".split()
            path = str(self.save_folder / f"{atom}-summary.tex")
            self.series_table(atom.lower(), bases, path)

        self.all_table(bases, str(self.save_folder / "all-summary.tex"))


class ExtrapSchemeSummary(CebeTable):
    col_names = ["Basis", "Method", "MSE", "MAE", "MedAE", "MaxAE", "STD"]

    def __init__(
        self,
        schemeToStats: SchemeStatsType,
        atomToSchemeStats: AtomStatsType,
        save_folder: Path,
        show_sample_size: bool = True,
        isPublication: bool = False,
    ):
        self.schemeToStats = schemeToStats
        self.atomToSchemeStats = atomToSchemeStats
        self.save_folder = save_folder
        self.show_sample_size = show_sample_size
        self.isPublication = isPublication

    def create_table_body(
        self, statsContainer: SchemeStatsType, schemes: Iterable[str]
    ) -> Tuple[List[List[str]], Set[int]]:
        body: List[List[str]] = []
        nSet: Set[int] = set()
        for scheme in schemes:
            stats = statsContainer[scheme]
            row = [scheme]
            for key in "MSE MAE MedAE MaxAE STD(AE)".split():
                row.append(self._format(cast(float, stats[cast(StatsKeyType, key)])))
            if self.show_sample_size:
                row.append(str(stats["n"]))
            nSet.add(int(stats["n"]))
            body.append(row)
        return body, nSet

    def series_table(self, atom: str, schemes: Iterable[str], save_path: str) -> None:
        body, nSet = self.create_table_body(self.atomToSchemeStats[atom], schemes)

        if self.show_sample_size:
            col_names = self.col_names + ["Sample Size"]
            suffix = ""
        else:
            col_names = self.col_names
            assert len(nSet) == 1, "Sample size must be the same for all methods"
            suffix = f" ({nSet.pop()} molecules)"
        header = ["\\textbf{%s}" % col_name for col_name in col_names]
        caption = (
            f"Statistical analysis of accuracy of different extrapolation schemes at predicting K-Edge CEBEs (in eV) compared to experimental data for {atom.upper()}-series"
            + suffix
        )
        latex.Table.export_table(
            caption=caption,
            label=f"extrap-scheme-summary-{atom.lower()}",
            positioning="l " * len(header),
            headers=header,
            body=body,
            save_path=save_path,
        )

    def all_table(self, schemes: Iterable[str], save_path: str) -> None:
        body, nSet = self.create_table_body(self.schemeToStats, schemes)

        if self.show_sample_size:
            self.col_names.append("Sample Size")
            suffix = ""
        else:
            assert len(nSet) == 1, "Sample size must be the same for all methods"
            suffix = f" ({nSet.pop()} molecules)"
        header = ["\\textbf{%s}" % col_name for col_name in self.col_names]
        caption = (
            "Statistical analysis of accuracy of different extrapolation schemes at predicting K-Edge CEBEs (in eV) compared to experimental data for all data points"
            + suffix
        )
        latex.Table.export_table(
            caption=caption,
            label="extrap-all-summary",
            positioning="l " * len(header),
            headers=header,
            body=body,
            save_path=save_path,
        )

    def results_for_schemes(self, scheme_factory: Callable[[], Iterable[str]]) -> None:
        atoms = "C N O F".split()
        for atom in atoms:
            path = str(self.save_folder / f"{atom}-summary.tex")
            self.series_table(atom.lower(), scheme_factory(), path)
        self.all_table(scheme_factory(), str(self.save_folder / "all-summary.tex"))


class UsedGeometries:

    def __init__(self, geom_wb: str, save_folder: Path) -> None:
        self.geom_wb = load_workbook(geom_wb)
        self.save_folder = save_folder

    def _get_formula(self, mol: str) -> str:
        return constants.FNAME_TO_MOLS[mol]["formula"]

    def _render_mol(self, mol: str) -> str:
        return constants.FNAME_TO_MOLS[mol]["latex"]

    def parse_geometries(self, relevantMols: Dict[str, Set[str]]) -> None:
        methodToAtomToMols: Dict[str, Dict[str, Dict[str, List[str]]]] = {}
        methodToBasisToMols: Dict[str, Dict[str, List[str]]] = {}
        ws = self.geom_wb["Sheet1"]
        row = 2
        while True:
            if (
                ws["B" + str(row)].value is None
                and ws["B" + str(row + 1)].value is None
            ):
                break
            if ws["B" + str(row)].value is None:
                row += 1
                continue
            atom = ws["B" + str(row)].value
            if atom not in relevantMols.keys():
                row += 1
                continue
            molecule = ws["C" + str(row)].value
            if molecule not in relevantMols[atom]:
                row += 1
                continue
            method = ws["D" + str(row)].value
            basis = (
                ws["E" + str(row)].value
                if ws["E" + str(row)].value is not None
                else "exp"
            )
            methodToAtomToMols.setdefault(method, {}).setdefault(atom, {}).setdefault(
                basis, []
            ).append(molecule)
            methodToBasisToMols.setdefault(method, {}).setdefault(basis, []).append(
                self._get_formula(molecule)
            )
            row += 1
        self.methodToAtomToBasisToMols = methodToAtomToMols
        self.methodToBasisToMols = methodToBasisToMols

    def summary(self, save_path: str) -> None:
        mols = list(sorted(self.methodToBasisToMols["exp"]["exp"]))
        self.methodToBasisToMols.setdefault("exp1", {})["exp1"] = mols[: len(mols) // 2]
        self.methodToBasisToMols.setdefault("exp2", {})["exp2"] = mols[len(mols) // 2 :]
        caption = "Molecules for which experimental geometries were available"
        header = ["exp1", "exp2", "MP2(Full)", "RI-MP2"]
        positioning = "p{0.30\linewidth} | p{0.30\linewidth} | p{0.30\linewidth} | p{0.30\linewidth}"
        body: List[List[str]] = []
        for j, method in enumerate(header):
            i = 0
            for basis in self.methodToBasisToMols[method]:
                if basis not in {"exp", "exp1", "exp2"}:
                    if i >= len(body):
                        body.append([])
                    body[i].append("\\textit{%s}" % basis)
                    i += 1
                for molecule in sorted(self.methodToBasisToMols[method][basis]):
                    cell = "\ch{%s}" % molecule
                    if i >= len(body):
                        body.append([])
                    body[i].append(cell)
                    i += 1

                cell = "(%s molecules)" % len(self.methodToBasisToMols[method][basis])
                if i >= len(body):
                    body.append([])
                body[i].append(cell)
                i += 1

            if j > 0:
                if i < len(body):
                    while i < len(body):
                        body[i].append(" ")
                        i += 1

        latex.Table.export_table(
            caption=caption,
            label="geom-summary",
            positioning=positioning,
            headers=header,
            body=body,
            save_path=save_path,
        )

    def main(self, relevantMols: Dict[str, Set[str]]) -> None:
        self.parse_geometries(relevantMols)
        self.summary(str(self.save_folder / "geom-summary.tex"))
